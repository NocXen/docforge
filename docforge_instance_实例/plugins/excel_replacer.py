"""
Excel表格替换插件
将数据填入Excel模板，替换所有字段

支持格式：xlsx, xls, xlsm, xlsb, odf, ods, odt

配置格式：
{
    "Target": ["template1.xlsx", "template2.xlsx"],
    "Ignore": ["temp.xlsx"],
    "FileType": "xlsx",
    "KeepPlaceHolderIfShorter": false,
    "ReplacerOfShorterPlaceHolder": ""
}

使用示例：
    # 在工作流中配置
    {
        "step_id": "replace",
        "plugin_name": "excel_replacer",
        "config": {
            "Target": ["数据表.xlsx"],
            "FileType": "xlsx"
        }
    }
"""

import re
from typing import Dict, List, Any, Optional
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import BasePlugin
from ..types import ExecutionResult
from ..constants import FileExtensions


class ExcelReplacer(BasePlugin):
    """
    Excel表格替换插件
    
    将数据填入Excel模板，替换所有字段
    支持：
    - 单元格中的字段替换
    - 多种输出格式
    - 继承模板的所有格式
    """
    
    def __init__(self):
        super().__init__()
        self._generated_names = set()
    
    @property
    def name(self) -> str:
        return "excel_replacer"
    
    @property
    def version(self) -> str:
        return "1.0.0"
    
    @property
    def plugin_type(self) -> str:
        return "replacer"
    
    @property
    def description(self) -> str:
        return "将数据填入Excel模板，替换所有字段"
    
    @property
    def dependencies(self) -> List[str]:
        return ["pandas", "openpyxl", "odfpy"]
    
    def execute(self, **kwargs) -> ExecutionResult:
        """
        执行Excel表格替换
        
        Args:
            input_files: 输入文件列表（模板文件）
            template_files: 模板文件列表
            output_dir: 输出目录
            data: 输入数据
            config: 配置参数
            plugin_api: 插件API
            
        Returns:
            ExecutionResult: 执行结果
        """
        try:
            # 获取参数
            template_files = kwargs.get("template_files", [])
            output_dir = kwargs.get("output_dir", "")
            data = kwargs.get("data", {})
            config = kwargs.get("config", {})
            plugin_api = kwargs.get("plugin_api")
            
            if not template_files:
                return ExecutionResult(
                    success=False,
                    errors=["没有模板文件"]
                )
            
            if not output_dir:
                return ExecutionResult(
                    success=False,
                    errors=["未设置输出目录"]
                )
            
            # 获取配置
            target = config.get("Target", [])
            ignore = config.get("Ignore", [])
            file_type = config.get("FileType", "xlsx")
            keep_placeholder_if_shorter = config.get("KeepPlaceHolderIfShorter", False)
            replacer_of_shorter_placeholder = config.get("ReplacerOfShorterPlaceHolder", "")
            
            # 验证输出格式
            if file_type not in ["xlsx", "xls", "xlsm", "xlsb", "odf", "ods", "odt"]:
                file_type = "xlsx"
            
            # 过滤模板文件
            filtered_templates = self._filter_templates(template_files, target, ignore)
            
            if not filtered_templates:
                return ExecutionResult(
                    success=False,
                    errors=["没有需要处理的模板文件"]
                )
            
            # 记录日志
            if plugin_api:
                plugin_api.log_info(f"开始替换Excel表格，共 {len(filtered_templates)} 个模板")
            
            # 处理每个模板
            output_files = []
            
            # 检查所有字段的值数量是否一致
            field_lengths = {key: len(values) for key, values in data.items() if isinstance(values, list)}
            if field_lengths:
                max_length = max(field_lengths.values())
                min_length = min(field_lengths.values())
                
                if max_length != min_length:
                    if plugin_api:
                        plugin_api.log_warning(f"字段值数量不一致: {field_lengths}，可能导致串值问题")
                
                # 遍历所有值的索引
                for index in range(max_length):
                    for template_path in filtered_templates:
                        try:
                            # 生成输出文件名（添加索引后缀）
                            output_name = self._generate_output_name(template_path, data, file_type, index)
                            output_path = Path(output_dir) / output_name
                            
                            # 替换表格（使用指定索引的值）
                            success = self._replace_workbook(
                                template_path,
                                output_path,
                                data,
                                file_type,
                                index,
                                keep_placeholder_if_shorter,
                                replacer_of_shorter_placeholder,
                                plugin_api
                            )
                            
                            if success:
                                output_files.append(str(output_path))
                                
                        except Exception as e:
                            if plugin_api:
                                plugin_api.log_error(f"处理模板失败: {template_path} - {e}")
                            continue
            else:
                # 如果没有字段数据，使用原来的逻辑
                for template_path in filtered_templates:
                    try:
                        # 生成输出文件名
                        output_name = self._generate_output_name(template_path, data, file_type)
                        output_path = Path(output_dir) / output_name
                        
                        # 替换表格
                        success = self._replace_workbook(
                            template_path,
                            output_path,
                            data,
                            file_type,
                            None,
                            keep_placeholder_if_shorter,
                            replacer_of_shorter_placeholder,
                            plugin_api
                        )
                        
                        if success:
                            output_files.append(str(output_path))
                            
                    except Exception as e:
                        if plugin_api:
                            plugin_api.log_error(f"处理模板失败: {template_path} - {e}")
                        continue
            
            if not output_files:
                return ExecutionResult(
                    success=False,
                    errors=["未能生成任何输出文件"]
                )
            
            # 记录日志
            if plugin_api:
                plugin_api.log_info(f"替换完成，生成 {len(output_files)} 个文件")
            
            return ExecutionResult(
                success=True,
                output_files=output_files
            )
            
        except Exception as e:
            return ExecutionResult(
                success=False,
                errors=[f"替换失败: {str(e)}"]
            )
    
    def _filter_templates(
        self,
        template_files: List[str],
        target: List[str],
        ignore: List[str]
    ) -> List[str]:
        """
        过滤模板文件
        
        Args:
            template_files: 所有模板文件
            target: 正选列表（只选择这些）
            ignore: 反选列表（不选择这些）
            
        Returns:
            List[str]: 过滤后的模板文件
        """
        result = []
        
        for template in template_files:
            template_name = Path(template).name
            
            # 检查是否应该忽略
            if ignore:
                should_ignore = False
                for ignore_pattern in ignore:
                    if self._match_pattern(template_name, ignore_pattern):
                        should_ignore = True
                        break
                if should_ignore:
                    continue
            
            # 检查是否应该选择
            if target:
                should_include = False
                for target_pattern in target:
                    if self._match_pattern(template_name, target_pattern):
                        should_include = True
                        break
                if not should_include:
                    continue
            
            result.append(template)
        
        return result
    
    def _match_pattern(self, filename: str, pattern: str) -> bool:
        """
        匹配文件名和模式
        
        支持：
        - 精确匹配
        - 通配符匹配（*）
        
        Args:
            filename: 文件名
            pattern: 匹配模式
            
        Returns:
            bool: 是否匹配
        """
        # 精确匹配
        if filename == pattern:
            return True
        
        # 通配符匹配
        if "*" in pattern:
            regex_pattern = pattern.replace("*", ".*")
            return bool(re.match(regex_pattern, filename))
        
        return False
    
    def _generate_output_name(self, template_path: str, data: Dict[str, List[str]], file_type: str, index: int = None) -> str:
        """
        生成输出文件名
        
        支持在文件名中使用字段，如：{公司名称}_数据表.xlsx
        如果生成的文件名重复，会自动添加后缀确保唯一性
        
        Args:
            template_path: 模板路径
            data: 数据字典
            file_type: 输出文件类型
            index: 值索引（用于多值情况）
            
        Returns:
            str: 输出文件名
        """
        template_name = Path(template_path).stem
        
        # 生成基础文件名
        if "{" in template_name:
            # 文件名包含字段，渲染后作为基础名称
            base_name = self._render_filename(template_name, data, index)
        else:
            # 文件名不包含字段，使用默认格式
            if index is not None:
                base_name = f"{template_name}_output_{index + 1}"
            else:
                base_name = f"{template_name}_output"
        
        # 生成完整文件名
        output_name = f"{base_name}.{file_type}"
        
        # 检查是否重名，如果重名则添加后缀
        counter = 1
        while output_name in self._generated_names:
            output_name = f"{base_name}_{counter}.{file_type}"
            counter += 1
        
        # 记录已生成的文件名
        self._generated_names.add(output_name)
        
        return output_name
    
    def _render_filename(self, filename: str, data: Dict[str, List[str]], index: int = None) -> str:
        """
        渲染文件名中的字段
        
        Args:
            filename: 文件名
            data: 数据字典
            index: 值索引（用于多值情况）
            
        Returns:
            str: 渲染后的文件名
        """
        def replacer(match):
            field_name = match.group(1)
            value = data.get(field_name, [field_name])
            
            if isinstance(value, list) and value:
                if index is not None and index < len(value):
                    return str(value[index])
                else:
                    return str(value[0])
            return str(value)
        
        pattern = r'\{([^}]+)\}'
        return re.sub(pattern, replacer, filename)
    
    def _replace_workbook(
        self,
        template_path: str,
        output_path: Path,
        data: Dict[str, List[str]],
        file_type: str,
        index: int,
        keep_placeholder_if_shorter: bool,
        replacer_of_shorter_placeholder: str,
        plugin_api=None
    ) -> bool:
        """
        替换工作簿中的字段
        
        Args:
            template_path: 模板路径
            output_path: 输出路径
            data: 数据字典
            file_type: 输出文件类型
            index: 值索引（用于多值情况）
            keep_placeholder_if_shorter: 是否保留短字段占位符
            replacer_of_shorter_placeholder: 短字段替换字符
            plugin_api: 插件API
            
        Returns:
            bool: 是否成功
        """
        try:
            # 加载模板
            template_full_path = Path(template_path)
            
            if not template_full_path.exists():
                if plugin_api:
                    plugin_api.log_error(f"模板文件不存在: {template_path}")
                return False
            
            # 检查文件格式
            template_ext = template_full_path.suffix.lower()
            
            # 对于xlsx格式，使用openpyxl保留格式
            if template_ext in [".xlsx", ".xlsm", ".xlsb"]:
                return self._replace_with_openpyxl(
                    template_full_path,
                    output_path,
                    data,
                    file_type,
                    index,
                    keep_placeholder_if_shorter,
                    replacer_of_shorter_placeholder,
                    plugin_api
                )
            else:
                # 对于其他格式，使用pandas
                return self._replace_with_pandas(
                    template_full_path,
                    output_path,
                    data,
                    file_type,
                    index,
                    keep_placeholder_if_shorter,
                    replacer_of_shorter_placeholder,
                    plugin_api
                )
            
        except Exception as e:
            if plugin_api:
                plugin_api.log_error(f"替换工作簿失败: {e}")
            return False
    
    def _replace_with_openpyxl(
        self,
        template_path: Path,
        output_path: Path,
        data: Dict[str, List[str]],
        file_type: str,
        index: int,
        keep_placeholder_if_shorter: bool,
        replacer_of_shorter_placeholder: str,
        plugin_api=None
    ) -> bool:
        """
        使用openpyxl替换（保留格式）
        
        Args:
            template_path: 模板路径
            output_path: 输出路径
            data: 数据字典
            file_type: 输出文件类型
            index: 值索引（用于多值情况）
            keep_placeholder_if_shorter: 是否保留短字段占位符
            replacer_of_shorter_placeholder: 短字段替换字符
            plugin_api: 插件API
            
        Returns:
            bool: 是否成功
        """
        try:
            # 加载工作簿
            wb = load_workbook(filename=template_path)
            
            # 获取数据（使用指定索引的值）
            replacements = {}
            for key, values in data.items():
                if isinstance(values, list) and values:
                    if index is not None and index < len(values):
                        replacements[key] = str(values[index])
                    else:
                        replacements[key] = str(values[0])
                else:
                    replacements[key] = str(values)
            
            # 按字段长度降序排序，避免短字段破坏长字段
            sorted_replacements = sorted(replacements.items(), key=lambda x: len(x[0]), reverse=True)
            
            # 遍历所有工作表
            for ws in wb.worksheets:
                # 遍历所有单元格
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original_text = cell.value
                            replaced_text = original_text
                            
                            for field, value in sorted_replacements:
                                if field in replaced_text:
                                    # 检查字段长度
                                    if len(value) < len(field) and keep_placeholder_if_shorter:
                                        # 保留占位符或使用自定义替换字符
                                        if replacer_of_shorter_placeholder:
                                            replaced_text = replaced_text.replace(field, replacer_of_shorter_placeholder)
                                        else:
                                            # 保留原占位符
                                            pass
                                    else:
                                        replaced_text = replaced_text.replace(field, value)
                            
                            if replaced_text != original_text:
                                cell.value = replaced_text
            
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存工作簿
            wb.save(output_path)
            
            if plugin_api:
                plugin_api.log_info(f"已生成: {output_path.name}")
            
            return True
            
        except Exception as e:
            if plugin_api:
                plugin_api.log_error(f"openpyxl替换失败: {e}")
            return False
    
    def _replace_with_pandas(
        self,
        template_path: Path,
        output_path: Path,
        data: Dict[str, List[str]],
        file_type: str,
        index: int,
        keep_placeholder_if_shorter: bool,
        replacer_of_shorter_placeholder: str,
        plugin_api=None
    ) -> bool:
        """
        使用pandas替换（通用格式）
        
        Args:
            template_path: 模板路径
            output_path: 输出路径
            data: 数据字典
            file_type: 输出文件类型
            index: 值索引（用于多值情况）
            keep_placeholder_if_shorter: 是否保留短字段占位符
            replacer_of_shorter_placeholder: 短字段替换字符
            plugin_api: 插件API
            
        Returns:
            bool: 是否成功
        """
        try:
            # 读取Excel文件
            excel_file = pd.ExcelFile(template_path)
            sheet_names = excel_file.sheet_names
            
            # 获取数据（使用指定索引的值）
            replacements = {}
            for key, values in data.items():
                if isinstance(values, list) and values:
                    if index is not None and index < len(values):
                        replacements[key] = str(values[index])
                    else:
                        replacements[key] = str(values[0])
                else:
                    replacements[key] = str(values)
            
            # 创建Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name in sheet_names:
                    # 读取工作表
                    df = pd.read_excel(template_path, sheet_name=sheet_name, header=None)
                    
                    # 替换字段
                    for col_idx in range(len(df.columns)):
                        for row_idx in range(len(df)):
                            cell_value = df.iloc[row_idx, col_idx]
                            
                            if pd.notna(cell_value) and isinstance(cell_value, str):
                                original_text = str(cell_value)
                                replaced_text = original_text
                                
                                for field, value in replacements.items():
                                    if field in replaced_text:
                                        # 检查字段长度
                                        if len(value) < len(field) and keep_placeholder_if_shorter:
                                            # 保留占位符或使用自定义替换字符
                                            if replacer_of_shorter_placeholder:
                                                replaced_text = replaced_text.replace(field, replacer_of_shorter_placeholder)
                                            else:
                                                # 保留原占位符
                                                pass
                                        else:
                                            replaced_text = replaced_text.replace(field, value)
                                
                                if replaced_text != original_text:
                                    df.iloc[row_idx, col_idx] = replaced_text
                    
                    # 写入工作表
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            if plugin_api:
                plugin_api.log_info(f"已生成: {output_path.name}")
            
            return True
            
        except Exception as e:
            if plugin_api:
                plugin_api.log_error(f"pandas替换失败: {e}")
            return False
    
    def cleanup(self) -> None:
        """
        清理资源
        
        在插件卸载或重新加载前调用，用于释放资源。
        Excel替换器没有需要清理的资源。
        """
        pass
